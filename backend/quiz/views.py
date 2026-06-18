import openpyxl
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q

from rest_framework import status, generics, views
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

from .models import User, Round, Question, Submission, Result, Winner, Notification
from .serializers import (
    UserSerializer, RegisterSerializer, RoundSerializer,
    QuestionStudentSerializer, QuestionAdminSerializer,
    SubmissionSerializer, SubmissionDetailSerializer,
    ResultSerializer, WinnerSerializer, NotificationSerializer
)
from .permissions import IsAdminUser

# Custom JWT Login View
class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        data = super().validate(attrs)
        data['user'] = UserSerializer(self.user).data
        return data

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

# Register API (Optional but extremely useful for setting up test users)
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]

# Current Round Status
class CurrentRoundView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rounds = Round.objects.all().order_by('id')
        user = request.user
        
        # Determine qualification for each round
        results = Result.objects.filter(user=user)
        results_map = {res.round_id: res for res in results}

        round_data = []
        for r in rounds:
            # Qualification logic:
            # Round 1 (first round in ID order): anyone can participate
            # Next rounds: qualified if they passed the previous round
            is_qualified = False
            prev_round = Round.objects.filter(id__lt=r.id).order_by('-id').first()
            if not prev_round:
                is_qualified = True  # Round 1 is open
            else:
                prev_res = results_map.get(prev_round.id)
                if prev_res and prev_res.qualified:
                    is_qualified = True

            user_result = results_map.get(r.id)
            round_data.append({
                'id': r.id,
                'name': r.name,
                'status': r.status,
                'start_time': r.start_time,
                'end_time': r.end_time,
                'results_published': r.results_published,
                'is_qualified': is_qualified,
                'qualified': user_result.qualified if user_result else False,
                'finalized': user_result.finalized if user_result else False,
                'score': user_result.score if user_result else 0
            })

        return Response({
            'user': UserSerializer(user).data,
            'rounds': round_data
        })

# Questions View (filters based on round and qualification)
class QuestionsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, round_id):
        round_obj = get_object_or_404(Round, id=round_id)
        user = request.user

        # 1. Check if round is active
        if round_obj.status != 'ACTIVE':
            return Response(
                {'error': f'{round_obj.name} is not active.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 2. Check if user is qualified for this round
        prev_round = Round.objects.filter(id__lt=round_obj.id).order_by('-id').first()
        if prev_round:
            prev_res = Result.objects.filter(user=user, round=prev_round).first()
            if not prev_res or not prev_res.qualified:
                return Response(
                    {'error': f'You are not qualified for {round_obj.name}.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # 3. Check if user already finalized this round
        user_res = Result.objects.filter(user=user, round=round_obj).first()
        if user_res and user_res.finalized:
            return Response(
                {'error': f'You have already submitted answers for {round_obj.name}.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 4. Return questions (no correct answers)
        questions = Question.objects.filter(round=round_obj).order_by('id')
        serializer = QuestionStudentSerializer(questions, many=True)

        # Also get any already saved drafts/submissions for this round
        submissions = Submission.objects.filter(user=user, question__round=round_obj)
        sub_data = {sub.question_id: sub.selected_answer for sub in submissions}

        return Response({
            'questions': serializer.data,
            'submissions': sub_data,
            'end_time': round_obj.end_time
        })

# Submit Answers View (supports draft and final submission)
class SubmitAnswersView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        round_id = request.data.get('round_id')
        answers = request.data.get('answers', [])  # list of {question_id: ..., selected_answer: ...}
        is_final = request.data.get('is_final', False)

        round_obj = get_object_or_404(Round, id=round_id)

        # Check if active
        if round_obj.status != 'ACTIVE':
            return Response({'error': 'Round is not active.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check qualification
        prev_round = Round.objects.filter(id__lt=round_obj.id).order_by('-id').first()
        if prev_round:
            prev_res = Result.objects.filter(user=user, round=prev_round).first()
            if not prev_res or not prev_res.qualified:
                return Response({'error': 'Not qualified for this round.'}, status=status.HTTP_403_FORBIDDEN)

        # Check if finalized
        res_obj, created = Result.objects.get_or_create(user=user, round=round_obj)
        if res_obj.finalized:
            return Response({'error': 'You have already finalized this round.'}, status=status.HTTP_400_BAD_REQUEST)

        # Save submissions
        for ans in answers:
            q_id = ans.get('question_id')
            sel_ans = ans.get('selected_answer')
            if q_id and sel_ans in ['A', 'B', 'C', 'D']:
                question_obj = get_object_or_404(Question, id=q_id, round=round_obj)
                Submission.objects.update_or_create(
                    user=user,
                    question=question_obj,
                    defaults={'selected_answer': sel_ans}
                )

        # If final submission, calculate score & freeze
        if is_final:
            questions = Question.objects.filter(round=round_obj)
            submissions = Submission.objects.filter(user=user, question__round=round_obj)
            sub_map = {sub.question_id: sub.selected_answer for sub in submissions}

            score = 0
            for q in questions:
                if sub_map.get(q.id) == q.correct_answer:
                    score += 1

            res_obj.score = score
            res_obj.finalized = True
            res_obj.save()

            return Response({
                'message': 'Quiz answers submitted and finalized successfully.',
                'score': score,
                'finalized': True
            })

        return Response({
            'message': 'Quiz progress saved.',
            'finalized': False
        })

# Results View
class UserResultsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        results = Result.objects.filter(user=user).order_by('round__id')
        serializer = ResultSerializer(results, many=True)
        
        # Add results_published info and hide scores for unpublished rounds
        results_data = []
        for item in serializer.data:
            round_obj = Round.objects.get(id=item['round'])
            item_data = dict(item)
            item_data['results_published'] = round_obj.results_published
            if not round_obj.results_published:
                # Hide score and qualification until results are published
                item_data['score'] = None
                item_data['qualified'] = None
            results_data.append(item_data)
        
        return Response(results_data)


# ==========================================
# ADMIN VIEWS
# ==========================================

class AdminStartRoundView(views.APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        round_id = request.data.get('round_id')
        duration_minutes = request.data.get('duration_minutes', 30)
        round_obj = get_object_or_404(Round, id=round_id)

        now = timezone.now()
        round_obj.status = 'ACTIVE'
        round_obj.start_time = now
        round_obj.end_time = now + timezone.timedelta(minutes=int(duration_minutes))
        round_obj.save()

        return Response({
            'message': f'{round_obj.name} started successfully.',
            'round': RoundSerializer(round_obj).data
        })

class AdminEndRoundView(views.APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        round_id = request.data.get('round_id')
        round_obj = get_object_or_404(Round, id=round_id)

        round_obj.status = 'COMPLETED'
        round_obj.save()

        # Auto-score any submissions that weren't finalized yet
        users = User.objects.filter(is_staff=False)
        questions = Question.objects.filter(round=round_obj)
        
        for user in users:
            # Check if this user had active submissions
            submissions = Submission.objects.filter(user=user, question__round=round_obj)
            if submissions.exists():
                res_obj, created = Result.objects.get_or_create(user=user, round=round_obj)
                if not res_obj.finalized:
                    sub_map = {sub.question_id: sub.selected_answer for sub in submissions}
                    score = sum(1 for q in questions if sub_map.get(q.id) == q.correct_answer)
                    res_obj.score = score
                    res_obj.finalized = True
                    res_obj.save()

        return Response({
            'message': f'{round_obj.name} ended. All pending submissions auto-finalized.',
            'round': RoundSerializer(round_obj).data
        })

# Question Management Views
class AdminQuestionListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = QuestionAdminSerializer

    def get_queryset(self):
        round_id = self.request.query_params.get('round_id')
        if round_id:
            return Question.objects.filter(round_id=round_id).order_by('id')
        return Question.objects.all().order_by('id')

class AdminQuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    queryset = Question.objects.all()
    serializer_class = QuestionAdminSerializer

# Submissions Management
class AdminSubmissionsListView(views.APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        round_id = request.query_params.get('round_id')
        search_query = request.query_params.get('search', '')

        submissions = Submission.objects.all()

        if round_id:
            submissions = submissions.filter(question__round_id=round_id)
        if search_query:
            submissions = submissions.filter(
                Q(user__name__icontains=search_query) |
                Q(user__register_no__icontains=search_query) |
                Q(user__department__icontains=search_query)
            )

        submissions = submissions.order_by('-submitted_at')
        serializer = SubmissionDetailSerializer(submissions, many=True)
        return Response(serializer.data)

# Publish Results & Selection System
class AdminPublishResultsView(views.APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        round_id = request.data.get('round_id')
        cutoff_score = request.data.get('cutoff_score')
        qualified_user_ids = request.data.get('qualified_user_ids') # None if not passed
        action = request.data.get('action', 'publish') # 'save' or 'publish'

        round_obj = get_object_or_404(Round, id=round_id)

        # Update qualifications if selection or cutoff is provided
        if qualified_user_ids is not None:
            # Qualify explicit list of user IDs
            qualified_set = set(int(uid) for uid in qualified_user_ids)
            all_users = User.objects.filter(is_staff=False)
            for user in all_users:
                if user.id in qualified_set:
                    res, created = Result.objects.get_or_create(
                        user=user,
                        round=round_obj,
                        defaults={'score': 0, 'finalized': True}
                    )
                    res.qualified = True
                    res.save()
                else:
                    res = Result.objects.filter(user=user, round=round_obj).first()
                    if res:
                        res.qualified = False
                        res.save()
        elif cutoff_score is not None:
            # Qualify based on score cutoff
            cutoff = int(cutoff_score)
            all_users = User.objects.filter(is_staff=False)
            for user in all_users:
                res = Result.objects.filter(user=user, round=round_obj).first()
                if res:
                    res.qualified = res.score >= cutoff
                    res.save()

        if action == 'publish':
            if round_obj.results_published:
                return Response({'message': f'Results for {round_obj.name} are already published.'}, status=status.HTTP_400_BAD_REQUEST)

            # Mark results as published
            round_obj.results_published = True
            round_obj.save()

            # Create notifications for all participants who have results
            results_for_round = Result.objects.filter(round=round_obj)
            for res in results_for_round:
                # Prevent duplicate notifications
                if not Notification.objects.filter(user=res.user, round=round_obj).exists():
                    if res.qualified:
                        Notification.objects.create(
                            user=res.user,
                            title=f'🎉 Congratulations! You qualified in {round_obj.name}',
                            message=f'You scored {res.score} marks in {round_obj.name} and have been selected for the next round. Keep up the great work!',
                            notification_type='QUALIFIED',
                            round=round_obj
                        )
                    else:
                        Notification.objects.create(
                            user=res.user,
                            title=f'Results Published for {round_obj.name}',
                            message=f'You scored {res.score} marks in {round_obj.name}. Unfortunately, you did not qualify for the next round. Thank you for participating!',
                            notification_type='NOT_QUALIFIED',
                            round=round_obj
                        )

            return Response({'message': f'Results for {round_obj.name} published successfully. Notifications sent to all participants.'})

        return Response({'message': 'Qualification selection saved successfully. Results are not published yet.'})

# Leaderboard Endpoint
class LeaderboardView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        round_id = request.query_params.get('round_id')
        if not round_id:
            return Response({'error': 'round_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        results = Result.objects.filter(round_id=round_id).order_by('-score', 'user__name')
        serializer = ResultSerializer(results, many=True)
        return Response(serializer.data)

# Winner Management
class WinnerView(views.APIView):
    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdminUser()]
        return [IsAuthenticated()]

    def get(self, request):
        winners = Winner.objects.all().order_by('position')
        serializer = WinnerSerializer(winners, many=True)
        return Response(serializer.data)

    def post(self, request):
        # Format expects: [{"position": 1, "user_id": 4}, ...]
        winners_data = request.data.get('winners', [])

        Winner.objects.all().delete() # Reset winners

        for item in winners_data:
            pos = item.get('position')
            u_id = item.get('user_id')
            if pos and u_id:
                user = get_object_or_404(User, id=u_id)
                Winner.objects.create(position=pos, user=user)

        winners = Winner.objects.all().order_by('position')
        serializer = WinnerSerializer(winners, many=True)
        return Response(serializer.data)

# Export Excel
class AdminExportExcelView(views.APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        wb = openpyxl.Workbook()

        # Sheet 1: Participants
        ws1 = wb.active
        ws1.title = "Participants"
        ws1.append(["User ID", "Name", "Register No", "Department", "Email"])
        for u in User.objects.filter(is_staff=False).order_by('register_no'):
            ws1.append([u.id, u.name, u.register_no, u.department, u.email])

        # Sheet 2: Results
        ws2 = wb.create_sheet(title="Results")
        ws2.append(["Result ID", "User ID", "Name", "Register No", "Round Name", "Score", "Qualified"])
        for r in Result.objects.all().order_by('round__id', '-score'):
            ws2.append([r.id, r.user.id, r.user.name, r.user.register_no, r.round.name, r.score, r.qualified])

        # Sheet 3: Submissions
        ws3 = wb.create_sheet(title="Submissions")
        ws3.append(["Submission ID", "User ID", "Name", "Register No", "Round Name", "Question", "Selected Answer", "Correct Answer", "Is Correct"])
        for s in Submission.objects.all().order_by('question__round__id', 'user__register_no'):
            is_correct = s.selected_answer == s.question.correct_answer
            ws3.append([
                s.id, s.user.id, s.user.name, s.user.register_no,
                s.question.round.name, s.question.question,
                s.selected_answer, s.question.correct_answer, is_correct
            ])

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Quiz_Competition_Data.xlsx"'
        return response


# Notification Views
class NotificationsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:50]
        serializer = NotificationSerializer(notifications, many=True)
        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
        return Response({
            'notifications': serializer.data,
            'unread_count': unread_count
        })

class MarkNotificationReadView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, notification_id):
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return Response({'message': 'Notification marked as read.'})

class MarkAllNotificationsReadView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'message': 'All notifications marked as read.'})

# Reset Competition
class AdminResetCompetitionView(views.APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        # 1. Delete Winners
        Winner.objects.all().delete()

        # 2. Delete Results
        Result.objects.all().delete()

        # 3. Delete Submissions
        Submission.objects.all().delete()

        # 4. Delete Notifications
        Notification.objects.all().delete()

        # 5. Reset Round status and results_published
        Round.objects.all().update(
            status='NOT_STARTED',
            start_time=None,
            end_time=None,
            results_published=False
        )

        return Response({'message': 'Competition successfully reset. All submissions, results, notifications, and winners cleared.'})
